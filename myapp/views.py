import io
import csv
import openpyxl
from django.shortcuts import render
from django.contrib import messages
from .models import *
from decimal import Decimal
# import pyrebase
# import os
# from firebase_admin import firestore





# config = {
#   apiKey: "AIzaSyAeEiACK0kkCDILhT_GeghKn1OfK-h8qHY",
#   authDomain: "secha-wear.firebaseapp.com",
#   projectId: "secha-wear",
#   storageBucket: "secha-wear.appspot.com",
#   messagingSenderId: "891457580339",
#   appId: "1:891457580339:web:c6a48d71e84e9c6fc96246",
#   measurementId: "G-E6NFQEG4SZ"
# };


# firebase = pyrebase.initialize_app(config)
# storage = firebase.storage()

def sales_views(files, request):
    if files.name.endswith(".csv"):
        data_set = files.read().decode('UTF-8')
        strings = io.StringIO(data_set)
        next(strings)

        for item in csv.reader(strings, delimiter=',', quotechar='|'):
            try:
                existing_item = Sales.objects.filter(ids=item[0]).first()
                if existing_item is None:
                    Sales.objects.create(
                        ids=item[0],
                        itemName=item[1],
                        quantity=item[2],
                        amount=item[3],
                        tax = item[4].value if item[4] is not None else 0,
                        tax_amount=item[5],
                        total_amount=item[6],
                    )
                else:
                    existing_item.ids = item[0]
                    existing_item.itemName = item[1]
                    existing_item.quantity = item[2]
                    existing_item.amount = item[3]
                    existing_item.tax = item[4] if item[4] is not None else 0
                    existing_item.tax_amount = item[5]
                    existing_item.total_amount = item[6]
                    existing_item.save()
            except Exception as e:
                pass

        Sales_Field.objects.create(file=files)
        messages.success(request, "File Uploaded Successfully at Sales")
    elif files.name.endswith(".xlsx"):
        wb = openpyxl.load_workbook(files)
        sheet = wb.active
            
        rows = list(sheet.iter_rows())
        total_rows = len(rows)
                
        for index, item in enumerate(rows[1:total_rows], start=2):  # Start from index 2 if there's a header row
            try:
                existing_item = Sales.objects.filter(ids=item[0].value).first()
                if existing_item is None:
                    Sales.objects.create(
                        ids=item[0].value,
                        itemName=item[1].value,
                        quantity=item[2].value,
                        amount=item[3].value,
                        tax = item[4].value if item[4] is not None else 0,
                        tax_amount=item[5].value,
                        total_amount=item[6].value,
                    )
                else:
                    existing_item.ids = item[0].value
                    existing_item.itemName = item[1].value
                    existing_item.quantity = item[2].value
                    existing_item.amount = item[3].value
                    existing_item.tax = item[4].value if (item[4].value) is not None else 0
                    existing_item.tax_amount = item[5].value
                    existing_item.total_amount = item[6].value
                    existing_item.save()
            except Exception as e:
                print(e)

        Sales_Field.objects.create(file=files)
        messages.success(request, "File Uploaded Successfully at Sales")
    else:
        messages.info(request, "Please upload an xlsx or csv file only")


def invoice_views(files, request):
    if files.name.endswith(".csv"):
        data_set = files.read().decode('UTF-8')
        strings = io.StringIO(data_set)
        next(strings)

        for item in csv.reader(strings, delimiter=',', quotechar='|'):
            try:
                existing_item = Invoice_Profit.objects.filter(ids=item[0]).first()
                if existing_item is None:
                    Invoice_Profit.objects.create(
                        ids=item[0],
                        date=item[1],
                        invoice_numb=item[2],
                        customer_name=item[3],
                        phone_numb=item[4],
                        amount=item[5],
                        cogs=item[6],
                        margin=item[7],
                    )
                else:
                    existing_item.ids = item[0]
                    existing_item.date = item[1]
                    existing_item.invoice_numb = item[2]
                    existing_item.customer_name = item[3]
                    existing_item.phone_numb = item[4]
                    existing_item.amount = item[5]
                    existing_item.cogs = item[6]
                    existing_item.margin = item[7]
                    existing_item.save()
            except Exception as e:
                pass

        Invoice_Profit_Field.objects.create(file=files)
        messages.success(request, "File Uploaded Successfully at Invoice")
    elif files.name.endswith(".xlsx"):
        wb = openpyxl.load_workbook(files)
        sheet = wb.active
        
        rows = list(sheet.iter_rows())
        total_rows = len(rows)
        
        for index, item in enumerate(rows[1:total_rows], start=2):  # Start from index 2 if there's a header row
            try:
                existing_item = Invoice_Profit.objects.filter(ids=item[0].value).first()
                if existing_item is None:
                    Invoice_Profit.objects.create(
                        ids=item[0].value,
                        date=item[1].value,
                        invoice_numb=item[2].value,
                        customer_name=item[3].value,
                        phone_numb=item[4].value,
                        amount=item[5].value,
                        cogs=item[6].value,
                        margin=item[7].value,
                    )
                else:
                    existing_item.ids = item[0].value
                    existing_item.date = item[1].value
                    existing_item.invoice_numb = item[2].value
                    existing_item.customer_name = item[3].value
                    existing_item.phone_numb = item[4].value
                    existing_item.amount = item[5].value
                    existing_item.cogs = item[6].value
                    existing_item.margin = item[7].value
                    existing_item.save()
            except Exception as e:
                pass

        Invoice_Profit_Field.objects.create(file=files)
        messages.success(request, "File Uploaded Successfully at Invoice")
    else:
        messages.info(request, "Please upload an xlsx or csv file only")


def current_stock(files, request):
    if files.name.endswith(".csv"):
        data_set = files.read().decode('UTF-8')
        strings = io.StringIO(data_set)
        next(strings)  # Skip header row

        for item in csv.reader(strings, delimiter=',', quotechar='|'):
            try:
                existing_item = Current_Stock.objects.filter(ids=item[0]).first()
                if existing_item is None:
                    Current_Stock.objects.create(
                        ids=item[0],
                        name=item[1],
                        itemCode=item[2],
                        stock=item[3],
                        amount=item[4],
                        price=int(item[4]) / int(item[3]) if int(item[3]) != 0 else 0  # Avoid division by zero
                    )
                else:
                    existing_item.ids = item[0]
                    existing_item.itemCode = item[2]
                    existing_item.stock = item[3]
                    existing_item.amount = item[4]
                    existing_item.price = int(item[4]) / int(item[3]) if int(item[3]) != 0 else 0
                    existing_item.save()
            except Exception as e:
                # Handle exceptions here
                pass

        Current_Stock_Field.objects.create(file=files)
        messages.success(request, "File Uploaded Successfully")
    elif files.name.endswith(".xlsx"):
        wb = openpyxl.load_workbook(files)
        sheet = wb.active

        rows = list(sheet.iter_rows())
        total_rows = len(rows)

        for index, item in enumerate(rows[1:total_rows], start=2):  # Start from index 2 if there's a header row
            try:
                existing_item = Current_Stock.objects.filter(ids=item[0].value).first()
                if existing_item is None:
                    Current_Stock.objects.create(
                        ids=item[0].value,
                        name=item[1].value,
                        itemCode=item[2].value,
                        stock=item[3].value,
                        amount=item[4].value,
                        price=int(item[4].value) / int(item[3].value) if int(item[3].value) != 0 else 0
                    )
                else:
                    existing_item.ids = item[0].value
                    existing_item.itemCode = item[2].value
                    existing_item.stock = item[3].value
                    existing_item.amount = item[4].value
                    existing_item.price = int(item[4].value) / int(item[3].value) if int(item[3].value) != 0 else 0
                    existing_item.save()
            except Exception as e:
                # Handle exceptions here
                pass

        Current_Stock_Field.objects.create(file=files)
        messages.success(request, "File Uploaded Successfully")
    else:
        messages.info(request, "Please upload an xlsx or csv file only")



def main_views(request):
    if request.method == 'POST':
        files = request.FILES['files']
        code = request.POST.get('code')
        selects = request.POST.get('selects')

        if code == "55555":
            if selects == "sales":
                sales_views(files, request)
            elif selects == "invoice":
                invoice_views(files, request)
            else:
                current_stock(files, request)
        else:
            messages.info(request, "Please enter correct code")

    return render(request, 'index.html')







def view_Current_Stock(request):
    item = Current_Stock.objects.all()
    total = 0
    stocks=0
    for i in Current_Stock.objects.all():
        total += int(i.amount)
        stocks += int(i.stock)
    
    data = {'item':item, 'total':total,'stocks':stocks}
    
    return render(request,'stock.html',data)


def view_Invoice(request):
    invoices = Invoice_Profit.objects.all()
    amts = Decimal('0.00')
    cogss = Decimal('0.00')
    margins = Decimal('0.00')

    for i in invoices:
        amts += Decimal(i.amount.replace(',', ''))
        cogss += Decimal(i.cogs.replace(',', ''))
        margins += Decimal(i.margin.replace(',', ''))

    data = {'invoices':invoices,'amts':amts,'cogss':cogss,'margins':margins}

    return render(request,'invoice.html',data)


def view_Sales(request):
    saless = Sales.objects.all()
    total_qty = 0
    amt = Decimal('0.00')
    total_amt = Decimal('0.00')

    for i in saless:
        total_qty += int(i.quantity)
        amt += Decimal(i.amount)
        total_amt += Decimal(i.total_amount.replace(',', ''))

    data = {'saless':saless,'total_qty':total_qty,'amt':amt,'total_amt':total_amt}

    return render(request,'sales.html',data)




